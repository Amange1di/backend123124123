from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import get_user_model
from django.db.models import Count, Q
from django.core.exceptions import ValidationError
import openpyxl
from io import BytesIO
from django.utils import timezone

from .models import User, Group, TeacherProfile, StudentProfile, BulkImport
from .serializers import (
    UserSerializer, UserCreateSerializer, UserUpdateSerializer,
    GroupSerializer, GroupDetailSerializer,
    TeacherProfileSerializer, TeacherProfileCreateSerializer,
    StudentProfileSerializer, StudentProfileCreateSerializer,
    BulkImportSerializer, BulkImportCreateSerializer, AuthResponseSerializer,
    CustomTokenObtainPairSerializer
)

UserModel = get_user_model()


class CustomTokenObtainPairView(TokenObtainPairView):
    """Кастомная аутентификация с ролью в токене"""
    serializer_class = CustomTokenObtainPairSerializer
    
    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        
        return Response({
            'access': data['access'],
            'refresh': data['refresh'],
            'user': data['user'],
            'role': data.get('role')
        })


class UserViewSet(viewsets.ModelViewSet):
    """ViewSet для управления пользователями"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        return UserSerializer
    
    def get_queryset(self):
        queryset = User.objects.all().select_related('teacher_profile', 'student_profile')
        role = self.request.query_params.get('role')
        group_id = self.request.query_params.get('group_id')
        
        if role:
            queryset = queryset.filter(role=role)
        if group_id:
            queryset = queryset.filter(role='student', student_groups__id=group_id).distinct()
        
        return queryset
    
    def get_permissions(self):
        # Разрешаем чтение себе для всех
        if self.action in ['me', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]
    
    def list(self, request, *args, **kwargs):
        # Список пользователей - только для админов
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Доступ только для администраторов'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().list(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        # Создание пользователей - только для админов
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Доступ только для администраторов'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        """Получить текущего пользователя"""
        serializer = UserSerializer(request.user)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Сброс пароля (только админ)"""
        if not request.user.is_admin:
            return Response(
                {'error': 'Только администратор может сбрасывать пароли'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        user = self.get_object()
        new_password = request.data.get('new_password')
        
        if not new_password or len(new_password) < 8:
            return Response(
                {'error': 'Пароль должен содержать минимум 8 символов'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.set_password(new_password)
        user.save()
        
        return Response({'message': 'Пароль успешно изменён'})
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Включить/выключить пользователя"""
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        
        return Response({
            'message': f'Пользователь {"активирован" if user.is_active else "деактивирован"}',
            'is_active': user.is_active
        })
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Статистика пользователей"""
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Доступно только администраторам'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        stats = {
            'total_users': User.objects.count(),
            'admins': User.objects.filter(role='admin').count(),
            'teachers': User.objects.filter(role='teacher').count(),
            'students': User.objects.filter(role='student').count(),
            'active_users': User.objects.filter(is_active=True).count(),
            'inactive_users': User.objects.filter(is_active=False).count(),
        }
        return Response(stats)


class GroupViewSet(viewsets.ModelViewSet):
    """ViewSet для управления группами"""
    queryset = Group.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return GroupDetailSerializer
        return GroupSerializer
    
    def get_queryset(self):
        queryset = Group.objects.prefetch_related('courses')
        
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def add_student(self, request, pk=None):
        """Добавить студента в группу"""
        group = self.get_object()
        student_id = request.data.get('student_id')
        
        try:
            student = User.objects.get(id=student_id, role='student')
            group.students.add(student)
            return Response({'message': 'Студент добавлен'})
        except User.DoesNotExist:
            return Response(
                {'error': 'Студент не найден'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def remove_student(self, request, pk=None):
        """Удалить студента из группы"""
        group = self.get_object()
        student_id = request.data.get('student_id')
        
        group.students.remove(student_id)
        return Response({'message': 'Студент удалён'})
    
    @action(detail=True, methods=['get'])
    def courses(self, request, pk=None):
        """Получить все курсы группы"""
        from courses.serializers import CourseSerializer
        group = self.get_object()
        courses = group.courses.all()
        serializer = CourseSerializer(courses, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Массовое создание групп из списка"""
        if not request.user.is_admin:
            return Response(
                {'error': 'Доступно только администраторам'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        groups_data = request.data.get('groups', [])
        created = []
        errors = []
        
        for group_data in groups_data:
            try:
                group = Group.objects.create(**group_data)
                created.append(group.group_name)
            except Exception as e:
                errors.append(f"{group_data.get('group_name', 'Unknown')}: {str(e)}")
        
        return Response({
            'created': created,
            'errors': errors,
            'total_created': len(created)
        })


class TeacherProfileViewSet(viewsets.ModelViewSet):
    """ViewSet для профилей преподавателей"""
    queryset = TeacherProfile.objects.all()
    serializer_class = TeacherProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'teacher':
            return TeacherProfile.objects.filter(user=user)
        return TeacherProfile.objects.all()
    
    @action(detail=False, methods=['get'])
    def my_profile(self, request):
        """Получить профиль текущего преподавателя"""
        try:
            profile = TeacherProfile.objects.get(user=request.user)
            return Response(TeacherProfileSerializer(profile).data)
        except TeacherProfile.DoesNotExist:
            return Response(
                {'error': 'Профиль не найден'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['put', 'patch'])
    def update_profile(self, request):
        """Обновить профиль преподавателя"""
        try:
            profile = TeacherProfile.objects.get(user=request.user)
            serializer = TeacherProfileSerializer(profile, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TeacherProfile.DoesNotExist:
            return Response(
                {'error': 'Профиль не найден'},
                status=status.HTTP_404_NOT_FOUND
            )


class StudentProfileViewSet(viewsets.ModelViewSet):
    """ViewSet для профилей студентов"""
    queryset = StudentProfile.objects.all()
    serializer_class = StudentProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'student':
            return StudentProfile.objects.filter(user=user)
        return StudentProfile.objects.all()
    
    @action(detail=False, methods=['get'])
    def my_profile(self, request):
        """Получить профиль текущего студента"""
        try:
            profile = StudentProfile.objects.get(user=request.user)
            return Response(StudentProfileSerializer(profile).data)
        except StudentProfile.DoesNotExist:
            return Response(
                {'error': 'Профиль не найден'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['put', 'patch'])
    def update_profile(self, request):
        """Обновить профиль студента"""
        try:
            profile = StudentProfile.objects.get(user=request.user)
            serializer = StudentProfileSerializer(profile, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except StudentProfile.DoesNotExist:
            return Response(
                {'error': 'Профиль не найден'},
                status=status.HTTP_404_NOT_FOUND
            )


class BulkImportViewSet(viewsets.ModelViewSet):
    """ViewSet для массового импорта данных"""
    queryset = BulkImport.objects.all()
    serializer_class = BulkImportSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return BulkImport.objects.all()
        return BulkImport.objects.none()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return BulkImportCreateSerializer
        return BulkImportSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def process(self, request, pk=None):
        """Обработать импорт"""
        if not request.user.is_admin:
            return Response(
                {'error': 'Доступно только администраторам'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        bulk_import = self.get_object()
        
        if bulk_import.status != 'pending':
            return Response(
                {'error': 'Импорт уже обрабатывается или обработан'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            bulk_import.status = 'processing'
            bulk_import.save()
            
            # Обработка Excel файла
            self._process_import(bulk_import)
            
        except Exception as e:
            bulk_import.status = 'failed'
            bulk_import.error_log = str(e)
            bulk_import.save()
            return Response(
                {'error': f'Ошибка обработки: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        return Response({
            'message': 'Импорт завершён',
            'successful': bulk_import.successful_rows,
            'failed': bulk_import.failed_rows
        })
    
    def _process_import(self, bulk_import):
        """Внутренний метод обработки импорта"""
        try:
            # Чтение Excel файла
            with bulk_import.file.open('rb') as f:
                wb = openpyxl.load_workbook(BytesIO(f.read()))
                ws = wb.active
                
                if bulk_import.import_type == 'students':
                    self._import_students(ws, bulk_import)
                elif bulk_import.import_type == 'teachers':
                    self._import_teachers(ws, bulk_import)
                elif bulk_import.import_type == 'groups':
                    self._import_groups(ws, bulk_import)
            
            bulk_import.status = 'completed'
            bulk_import.processed_at = timezone.now()
            bulk_import.save()
            
        except Exception as e:
            bulk_import.status = 'failed'
            bulk_import.error_log = str(e)
            bulk_import.save()
            raise
    
    def _import_students(self, ws, bulk_import):
        """Импорт студентов"""
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                username = row[0].value
                email = row[1].value
                first_name = row[2].value
                last_name = row[3].value
                group_name = row[4].value
                
                if not all([username, email]):
                    raise ValidationError("Не заполнены обязательные поля")
                
                # Создание пользователя
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password='TempPass123!',
                    first_name=first_name or '',
                    last_name=last_name or '',
                    role='student'
                )
                
                # Добавление в группу
                if group_name:
                    try:
                        group = Group.objects.get(group_name=group_name)
                        group.students.add(user)
                    except Group.DoesNotExist:
                        pass
                
                bulk_import.successful_rows += 1
            except Exception as e:
                bulk_import.failed_rows += 1
                bulk_import.error_log += f"Строка {row_idx}: {str(e)}\n"
            
            bulk_import.total_rows = row_idx - 1
        
        bulk_import.save()
    
    def _import_teachers(self, ws, bulk_import):
        """Импорт преподавателей"""
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                username = row[0].value
                email = row[1].value
                first_name = row[2].value
                last_name = row[3].value
                department = row[4].value
                position = row[5].value
                
                if not all([username, email]):
                    raise ValidationError("Не заполнены обязательные поля")
                
                # Создание пользователя
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password='TempPass123!',
                    first_name=first_name or '',
                    last_name=last_name or '',
                    role='teacher'
                )
                
                # Создание профиля
                TeacherProfile.objects.create(
                    user=user,
                    department=department or 'Общая',
                    position=position or 'Преподаватель'
                )
                
                bulk_import.successful_rows += 1
            except Exception as e:
                bulk_import.failed_rows += 1
                bulk_import.error_log += f"Строка {row_idx}: {str(e)}\n"
            
            bulk_import.total_rows = row_idx - 1
        
        bulk_import.save()
    
    def _import_groups(self, ws, bulk_import):
        """Импорт групп"""
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                group_name = row[0].value
                department = row[1].value
                semester = row[2].value
                year = row[3].value
                
                if not group_name:
                    raise ValidationError("Не заполнено название группы")
                
                Group.objects.create(
                    group_name=group_name,
                    department=department or '',
                    semester=semester or 1,
                    year=year or 2024
                )
                
                bulk_import.successful_rows += 1
            except Exception as e:
                bulk_import.failed_rows += 1
                bulk_import.error_log += f"Строка {row_idx}: {str(e)}\n"
            
            bulk_import.total_rows = row_idx - 1
        
        bulk_import.save()